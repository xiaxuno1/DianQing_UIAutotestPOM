# --------------------------------------------------
# !/usr/bin/python
# -*- coding: utf-8 -*-
# PN: DianQing_UIAutoTest
# FN: excel_conf
# Author: xiaxu
# DATA: 2022/7/21
# Description:设置excel的单元格格式
# ---------------------------------------------------
from openpyxl.styles import PatternFill,Font

def write_result(cell,row,columm,status=1):
    if status ==1:
        cell(row=row,columm=columm).value = "Pass"
        cell(row=row, columm=columm).fill = PatternFill("Solid",fgColor="AACF91")
    elif status == 2:
        cell(row=row, columm=columm).value = "Fail"
        cell(row=row, columm=columm).fill = PatternFill("Solid", fgColor="FF0000")
    else:
        cell(row=row, columm=columm).value = "Blcoked"
        cell(row=row, columm=columm).fill = PatternFill("Solid", fgColor="CDCDCD")#浅灰色
